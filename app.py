from email.message import EmailMessage
import json
import shutil
import smtplib

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from datetime import datetime
import win32print
from pydantic import BaseModel
from openpyxl import load_workbook
import uvicorn
import os
from typing import List, Optional
from fastapi.middleware.cors import CORSMiddleware
import threading
import time

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"], # Permite que o front-end acesse a API
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- CONFIGURAÇÕES DE CAMINHOS ---
PATH_BASE = r"C:\Users\smicael\ACHE LABORATORIOS FARMACEUTICOS SA\Sharepoint l Operações Logísticas - Carga direta\Gestão Entrega direta_Nova Versão_ 2025.xlsm"
PATH_BASE_2 = r"C:\Users\smicael\ACHE LABORATORIOS FARMACEUTICOS SA\Sharepoint l Operações Logísticas - Last Mile\LAST MILE - PL-PE-LOG-00031.xlsm" 
PATH_RELATORIO = r"C:\Users\smicael\ACHE LABORATORIOS FARMACEUTICOS SA\Sharepoint l Operações Logísticas - Carga direta\Base de dados\Relatorio_Bipagens.xlsx"
NOME_ABA = "NotasFiscais" 

# --- CARREGAMENTO E ATUALIZAÇÃO AUTOMÁTICA ---
print("Iniciando sistema de sincronização de bases...")
df_memoria_1 = pd.DataFrame()
df_memoria_2 = pd.DataFrame()

def atualizar_bases_loop():
    global df_memoria_1, df_memoria_2
    while True:
        try:
            # Carregamento Silencioso (evita poluir o terminal a cada 30s se quiser)
            df_temp1 = pd.read_excel(PATH_BASE, engine='openpyxl', sheet_name=NOME_ABA)
            df_temp2 = pd.read_excel(PATH_BASE_2, engine='openpyxl', sheet_name=NOME_ABA)
            
            # Atualiza as globais
            df_memoria_1 = df_temp1
            df_memoria_2 = df_temp2
            print(f"[{datetime.now().strftime('%H:%M:%S')}] Bases Excel atualizadas com sucesso.")
        except Exception as e:
            print(f"Erro ao atualizar bases: {e}")
        
        time.sleep(30) # Aguarda 30 segundos

# Inicia a atualização em uma Thread separada para não travar a API
threading.Thread(target=atualizar_bases_loop, daemon=True).start()

try:
    df_memoria_1 = pd.read_excel(PATH_BASE, engine='openpyxl', sheet_name=NOME_ABA)
    df_memoria_2 = pd.read_excel(PATH_BASE_2, engine='openpyxl', sheet_name=NOME_ABA)
    print("Bases carregadas com sucesso!")
except Exception as e:
    print(f"Erro ao carregar bases: {e}")

# --- MODELOS DE ENTRADA ---
class Armazenamento(BaseModel):
    id_pallet: str
    endereco: str

class Expedicao(BaseModel):
    id_pallet: str
    doca: str

class DadosBip(BaseModel):
    codigo: str
    tipo_processo: str

class PalletInfo(BaseModel):
    id_pallet: str
    carga: str
    cliente: str = "-"
    destino: str = "-"
    quantidade: int = 0
    qtd_remessas: int = 0
    remessas_lista: str = ""

class LotePallets(BaseModel):
    pallets: List[PalletInfo]

# --- FUNÇÃO DE IMPRESSÃO ZEBRA (ATUALIZADA E ATIVA) ---
def enviar_para_zebra(zpl):
    try:
        nome_zebra = None
        impressoras = win32print.EnumPrinters(win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS)
        for p in impressoras:
            # Busca melhorada por drivers comuns da Zebra
            if any(x in p[2].upper() for x in ["ZEBRA", "ZDESIGNER", "ZD", "GK420"]):
                nome_zebra = p[2]
                break
        
        if not nome_zebra:
            print("Impressora Zebra não encontrada!")
            return False

        hPrinter = win32print.OpenPrinter(nome_zebra)
        try:
            win32print.StartDocPrinter(hPrinter, 1, ("Etiqueta Pallet MC TECH", None, "RAW"))
            win32print.StartPagePrinter(hPrinter)
            win32print.WritePrinter(hPrinter, zpl.encode('utf-8'))
            win32print.EndPagePrinter(hPrinter)
            win32print.EndDocPrinter(hPrinter)
            return True
        finally:
            win32print.ClosePrinter(hPrinter)
    except Exception as e:
        print(f"Erro na impressão: {e}")
        return False

# --- FUNÇÕES AUXILIARES ---
def buscar_no_cache(df_origem, remessa_extraida):
    if df_origem.empty: return None
    cols_norm = {str(c).strip().lower(): c for c in df_origem.columns}
    col_ref = cols_norm.get("remessa")
    if col_ref:
        res_busca = df_origem[df_origem[col_ref].astype(str).str.replace(r'\.0$', '', regex=True).str.strip() == remessa_extraida]
        if not res_busca.empty: return res_busca.iloc[0], cols_norm
    if len(df_origem.columns) >= 9:
        col_9 = df_origem.columns[8]
        res_busca = df_origem[df_origem[col_9].astype(str).str.replace(r'\.0$', '', regex=True).str.strip() == remessa_extraida]
        if not res_busca.empty: return res_busca.iloc[0], cols_norm
    return None

def get_val(res, cols_norm, possibilidades):
    for p in possibilidades:
        p_lower = p.lower()
        if p_lower in cols_norm:
            val = res[cols_norm[p_lower]]
            return str(val).split('.')[0] if pd.notna(val) else "-"
    return "-"

# --- ROTAS PRINCIPAIS ---

@app.post("/registrar_bip")
async def registrar_bip(d: DadosBip):
    try:
        codigo_limpo = str(d.codigo).strip()
        if len(codigo_limpo) == 31:
            remessa_extraida = codigo_limpo[12:20].strip()
        else:
            remessa_extraida = codigo_limpo[10:18].strip() if len(codigo_limpo) >= 18 else "S/R"

        resultado = buscar_no_cache(df_memoria_1, remessa_extraida) or buscar_no_cache(df_memoria_2, remessa_extraida)

        if not resultado:
            return {"status": "erro", "dados": {"status": "REMESSA NÃO LOCALIZADA", "valido": False}}

        res, cols_norm = resultado
        carga_val = get_val(res, cols_norm, ["Carga", "Num Carga"])
        nf_val = get_val(res, cols_norm, ["NF-e", "NF", "Nota Fiscal"])
        cliente_val = get_val(res, cols_norm, ["Nome", "Cliente", "Razão Social"])
        cidade_val = get_val(res, cols_norm, ["Cidade"])
        regiao_val = get_val(res, cols_norm, ["Região", "Regiao"])
        
        try: 
            c_vol = cols_norm.get("vol")
            vol_total_remessa = int(float(res[c_vol])) if c_vol and pd.notna(res[c_vol]) else 1
        except: vol_total_remessa = 1

        if not os.path.exists(PATH_RELATORIO):
            pd.DataFrame(columns=["Codigo","Lote","Remessa","VolumeCli","VolumePr","Status","DataHora","Carga","Cliente","NF","Tipo","Cidade","Regiao","Paletizacao","Endereco","Doca"]).to_excel(PATH_RELATORIO, index=False)
        
        df_rel = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        ja_bipado = df_rel[(df_rel.iloc[:, 0] == codigo_limpo) & (df_rel.iloc[:, 5] == "✔ Conferido")]
        vol_lidos = len(df_rel[(df_rel.iloc[:, 2] == remessa_extraida) & (df_rel.iloc[:, 5] == "✔ Conferido")])

        info_tela = {
            "valido": True, "status": "CONFERIDO", "nf": nf_val,
            "vol_bipados": vol_lidos + (0 if not ja_bipado.empty else 1),
            "vol_total": vol_total_remessa, "carga_planilha": carga_val,
            "cliente_planilha": cliente_val, "cidade": cidade_val, "regiao": regiao_val
        }

        if not ja_bipado.empty:
            info_tela["status"] = "JÁ BIPADO"
            return {"status": "duplicado", "remessa": remessa_extraida, "dados": info_tela}

        wb = load_workbook(PATH_RELATORIO)
        ws = wb.active
        proxima_linha = 2
        while ws.cell(row=proxima_linha, column=1).value is not None:
            proxima_linha += 1
            
        agora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        dados_linha = [codigo_limpo, "-", remessa_extraida, "-", "-", "✔ Conferido", agora, carga_val, cliente_val, nf_val, d.tipo_processo, cidade_val, regiao_val, ""]
        
        for col, valor in enumerate(dados_linha, start=1):
            ws.cell(row=proxima_linha, column=col).value = valor
        wb.save(PATH_RELATORIO)
        return {"status": "sucesso", "remessa": remessa_extraida, "dados": info_tela}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/pallets-da-carga/{num_carga}")
async def pallets_da_carga(num_carga: str):
    try:
        if not os.path.exists(PATH_RELATORIO): return {"pallets": []}
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        df.columns = [str(c).strip() for c in df.columns]
        
        filtro = df[df['Carga'] == str(num_carga).strip()]
        if filtro.empty: return {"pallets": []}

        resultado = []
        ids_unicos = filtro[filtro['Paletizacao'].notna()].get('Paletizacao', pd.Series()).unique()
        
        for id_plt in ids_unicos:
            for id_plt in ids_unicos:
                if not id_plt or str(id_plt) == 'nan': continue
            
            dados_plt = filtro[filtro['Paletizacao'] == id_plt]
            
            # --- NOVA LÓGICA DE FORMATAÇÃO: REMESSA + NF (000000000-100) ---
            lista_combinada = []
            # Remove duplicatas baseadas no par Remessa/NF para não repetir na etiqueta
            pares_unicos = dados_plt.drop_duplicates(subset=['Remessa', 'NF'])
            
            for _, row in pares_unicos.iterrows():
                rem = str(row['Remessa']).strip()
                nf_bruta = str(row['NF']).strip()
                
                # Formata a NF: preenche com zeros até 9 dígitos e adiciona -100
                # Exemplo: '5521' vira '000005521-100'
                nf_formatada = nf_bruta.zfill(9)
                
                # Adiciona à lista com um separador visual claro
                lista_combinada.append(f"R:{rem} | NF:{nf_formatada}")
            
            # Une os itens com uma quebra de linha (representada por \, no ZPL) ou vírgula
            notas_e_remessas = ",  ".join(lista_combinada)
            # --------------------------------------------------------------

            # Busca a região (mantendo sua lógica anterior)
            regiao = "-"
            if 'Regiao' in dados_plt.columns:
                regiao = str(dados_plt['Regiao'].iloc[0]).strip().upper()
            elif 'Região' in dados_plt.columns:
                regiao = str(dados_plt['Região'].iloc[0]).strip().upper()
            else:
                regiao = str(dados_plt.iloc[0, 12]).strip().upper()

            resultado.append({
                "id_pallet": id_plt,
                "carga": num_carga,
                "cliente": ", ".join(filter(None, dados_plt['Cliente'].unique().astype(str))),
                "destino": regiao if regiao != "NAN" else "-",
                "quantidade": len(dados_plt),
                "qtd_remessas": dados_plt['Remessa'].nunique(),
                "remessas_lista": notas_e_remessas 
            })
        return {"pallets": resultado}
    except Exception as e:
        print(f"Erro ao buscar pallets: {e}")
        return {"pallets": []}

@app.post("/imprimir-todos-pallets")
async def imprimir_todos_pallets(lote: LotePallets):
    try:
        contagem = 0
        agora = datetime.now().strftime('%d/%m/%y %H:%M')
        
        for p in lote.pallets:
            regiao_upper = str(p.destino).upper().strip()
            titulo_cabecalho = "LAST MILE" if regiao_upper in ["PB", "PE"] else "CARGA DIRETA"
# --- LÓGICA DE SEPARAÇÃO PARA A ETIQUETA ---
            # Vamos separar a string p.remessas_lista que vem no formato "R:123 | NF:000-100, R:124 | NF:001-100"
            itens = [item.strip() for item in p.remessas_lista.split(',')]
            lista_rems = []
            lista_nfs = []
            
            for i in itens:
                if "|" in i:
                    partes = i.split("|")
                    lista_rems.append(partes[0].replace("R:", "").strip())
                    lista_nfs.append(partes[1].replace("NF:", "").strip())
            
            rems_str = ", ".join(lista_rems)
            nfs_str = ", ".join(lista_nfs)

            zpl = (
                f"^XA^CI28^PW800^LL1200^LS0"
                
                # Cabeçalho e Região
                f"^CF0,60^FO50,40^FB700,1,0,C^FD{titulo_cabecalho}^FS" 
                f"^FO50,105^GB700,3,3^FS"
                f"^CF0,30^FO50,140^FDREGIÃO:^FS^CF0,40^FO170,135^FD{regiao_upper}^FS"
                f"^CF0,35^FO50,230^FDCARGA: {p.carga}^FS"
                f"^CF0,35^FO450,230^FDVOLUMES: {p.quantidade}^FS"
                
                # Código de Barras Centralizado
                f"^BY2,3,100^FO80,300^BCN,80,Y,N,N^FD{p.id_pallet}^FS"
                
                # --- QUADRO GERAL DE DETALHES ---
                f"^FO50,450^GB700,630,3^FS" 
                
                # SEÇÃO 1: REMESSAS
                f"^CF0,30^FO70,470^FDLISTA DE REMESSAS^FS"
                f"^FO70,505^GB660,1,1^FS"
                f"^CF0,25^FO70,520^FB650,8,0,L^FD{rems_str}^FS" 
                
                # LINHA DIVISÓRIA CENTRAL
                f"^FO70,750^GB660,2,2^FS"
                
                # SEÇÃO 2: NOTAS FISCAIS
                f"^CF0,30^FO70,770^FDLISTA DE NOTAS FISCAIS^FS"
                f"^FO70,805^GB660,1,1^FS"
                f"^CF0,25^FO70,820^FB650,12,0,L^FD{nfs_str}^FS" 
                
                # Rodapé
                f"^CF0,20^FO50,1150^FB700,1,0,R^FD{agora} | {titulo_cabecalho}^FS"
                f"^XZ"
            )
            
            if enviar_para_zebra(zpl):
                contagem += 1
                
        return {"status": "ok", "impressos": contagem}
    except Exception as e:
        return {"status": "erro", "detalhe": str(e)}

@app.post("/armazenar-pallet")
async def armazenar_pallet(a: Armazenamento):
    try:
        wb = load_workbook(PATH_RELATORIO)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[13].value) == str(a.id_pallet):
                row[14].value = a.endereco
        wb.save(PATH_RELATORIO)
        return {"status": "ok"}
    except Exception as e: raise HTTPException(status_code=500, detail=str(e))

@app.post("/expedir-pallet")
async def expedir_pallet(e: Expedicao):
    try:
        wb = load_workbook(PATH_RELATORIO)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[13].value) == str(e.id_pallet):
                row[16].value = e.doca 
        wb.save(PATH_RELATORIO)
        return {"status": "OK"}
    except Exception as err: raise HTTPException(status_code=500, detail=str(err))

@app.post("/fechar-pallet")
async def fechar_pallet(p: PalletInfo):
    try:
        if not os.path.exists(PATH_RELATORIO):
            raise HTTPException(status_code=404, detail="Relatório não encontrado")

        wb = load_workbook(PATH_RELATORIO)
        ws = wb.active
        
        # 1. Preparar a data com pontos (Ex: 25.03.2026)
        data_atual = datetime.now().strftime("%d.%m.%Y")
        prefixo_busca = f"PLT{p.carga} - {data_atual}"
        
        # 2. Ler o Excel para encontrar a última sequência
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        
        # Filtra IDs que contenham o prefixo (Carga + Data)
        pallets_hoje = df[df.iloc[:, 13].astype(str).str.contains(prefixo_busca, na=False)]
        ids_existentes = pallets_hoje.iloc[:, 13].unique()
        
        proxima_seq = 1
        if len(ids_existentes) > 0:
            sequencias = []
            for id_plt in ids_existentes:
                try:
                    # Pega a parte após o último " - " (a sequência 0001)
                    partes = str(id_plt).split(' - ')
                    if len(partes) >= 3:
                        sequencias.append(int(partes[2]))
                except:
                    continue
            if sequencias:
                proxima_seq = max(sequencias) + 1

        # 3. Montar o ID final: PLT506 - 25.03.2026 - 0001
        id_final_formatado = f"PLT{p.carga} - {data_atual} - {str(proxima_seq).zfill(4)}"
        
        # 4. Gravar no Excel (Coluna 14 / Índice 13)
        gravou = False
        for row in ws.iter_rows(min_row=2):
            val_carga = str(row[7].value).strip()
            val_pallet = str(row[13].value).strip() if row[13].value else ""
            
            if val_carga == str(p.carga).strip() and (val_pallet == "" or val_pallet.lower() == "nan"):
                row[13].value = id_final_formatado
                gravou = True
        
        if not gravou:
            return {"status": "erro", "mensagem": "Nenhum volume pendente nesta carga."}

        wb.save(PATH_RELATORIO)
        return {"status": "ok", "id_pallet": id_final_formatado}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/produtividade")
async def produtividade():
    try:
        if not os.path.exists(PATH_RELATORIO): return {"total": 0}
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl')
        if df.empty: return {"total": 0}
        df['DataHora_DT'] = pd.to_datetime(df['DataHora'], dayfirst=True, errors='coerce')
        df_hoje = df[df['DataHora_DT'].dt.date == datetime.now().date()]
        return {"total": len(df_hoje)}
    except: return {"total": 0}

@app.get("/posicoes-ocupadas")
async def posicoes_ocupadas():
    try:
        if not os.path.exists(PATH_RELATORIO): return []
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        
        # Filtro: Tem endereço (índice 14) E a Expedição (índice 15) NÃO é "OK"
        df_vivos = df[
            (df.iloc[:, 14].notna()) & 
            (df.iloc[:, 14].str.strip() != "") & 
            (df.iloc[:, 15].astype(str).str.upper() != "OK")
        ]
        
        ocupados = df_vivos.iloc[:, 14].unique().tolist()
        return [str(x).strip() for x in ocupados if str(x).strip() not in ["", "nan"]]
    except Exception as e:
        print(f"Erro ao buscar ocupados: {e}")
        return []

@app.get("/mapa-por-carga")
async def mapa_por_carga():
    try:
        if not os.path.exists(PATH_RELATORIO): return {}
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        
        # Filtro: Tem endereço E Expedição NÃO é "OK"
        df_v = df[
            (df.iloc[:, 14].notna()) & 
            (df.iloc[:, 14].str.strip() != "") & 
            (df.iloc[:, 15].astype(str).str.upper() != "OK")
        ]
        
        res = {}
        for carga, grupo in df_v.groupby(df_v.columns[7]):
            res[str(carga).strip()] = grupo.iloc[:, 14].unique().tolist()
        return res
    except: return {}

@app.get("/detalhes-posicao/{endereco}")
async def detalhes_posicao(endereco: str):
    try:
        if not os.path.exists(PATH_RELATORIO): 
            return {"status": "vazio"}
            
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        
        # Filtra pelo endereço E garante que a Expedição (índice 15) não seja OK
        filtro = df[
            (df.iloc[:, 14] == endereco) & 
            (df.iloc[:, 15].astype(str).str.upper() != "OK")
        ]
        
        if filtro.empty: 
            return {"status": "vazio"}
            
        # --- NOVA LÓGICA PARA LISTAS ---
        # Pegamos todos os clientes únicos e removemos eventuais 'nan'
        lista_clientes = [c for c in filtro.iloc[:, 8].unique().tolist() if str(c).lower() != 'nan']
        
        # Pegamos todas as remessas únicas (coluna índice 2)
        lista_rems = [r for r in filtro.iloc[:, 2].unique().tolist() if str(r).lower() != 'nan']
        
        return {
            "status": "ocupado",
            "id_pallet": str(filtro.iloc[0, 13]),
            "carga": str(filtro.iloc[0, 7]),
            "clientes": ", ".join(lista_clientes), # Transforma a lista em texto separado por vírgula
            "total_remessas": len(lista_rems),
            "lista_remessas": ", ".join(lista_rems) # Transforma a lista em texto para o alert
        }
    except Exception as e:
        print(f"Erro detalhes posição: {e}")
        return {"status": "erro"}

@app.get("/detalhes-pallet-especifico/{id_pallet}")
async def detalhes_pallet_especifico(id_pallet: str):
    try:
        if not os.path.exists(PATH_RELATORIO): 
            raise HTTPException(status_code=404, detail="Relatório não encontrado")
            
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        
        # Filtra pelo ID do Pallet na coluna de índice 13 (Paletizacao)
        filtro = df[df.iloc[:, 13].astype(str).str.strip() == str(id_pallet).strip()].copy()
        
        if filtro.empty:
            return {"status": "vazio", "itens": [], "id_pallet": id_pallet}

        # --- LÓGICA DE AGRUPAMENTO E SOMA ---
        # 1. Identifica os índices das colunas para evitar erro de nome
        # Remessa: 2 | Carga: 7 | NF: 9 | Volume: Vamos assumir que é a 3 (VolumeCli/Pr) ou conte as bips
        # Como seu relatório é de bips, cada linha é 1 volume. Vamos contar as ocorrências.
        
        # Agrupamos por NF (9) e Remessa (2) e contamos quantas linhas existem para cada par
        agrupado = filtro.groupby([filtro.iloc[:, 9], filtro.iloc[:, 2]]).size().reset_index(name='volume_total')
        
        # Identifica o Tipo de Operação (Baseado na coluna 12 - Região/Destino)
        regiao = str(filtro.iloc[0, 12]).upper()
        tipo_op = "LAST MILE" if any(x in regiao for x in ["PB", "PE"]) else "CARGA DIRETA"

        itens = []
        for _, row in agrupado.iterrows():
            nf_original = str(row.iloc[0]).strip() # Índice 0 do agrupado é a NF
            remessa = str(row.iloc[1]).strip()     # Índice 1 do agrupado é a Remessa
            soma_volumes = str(row['volume_total']) # Resultado do size()
            
            nf_formatada = nf_original.zfill(9) + "-100" if nf_original != "nan" else "-"
            
            itens.append({
                "nf": nf_formatada,
                "remessa": remessa,
                "volume": soma_volumes 
            })

        return {
            "id_pallet": id_pallet,
            "carga": str(filtro.iloc[0, 7]).strip(),
            "tipo_operacao": tipo_op,
            "itens": itens
        }
    except Exception as e:
        print(f"Erro no agrupamento: {e}")
        raise HTTPException(status_code=500, detail=str(e))
    
@app.get("/ids-para-espelho")
async def ids_para_espelho():
    try:
        if not os.path.exists(PATH_RELATORIO): return []
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        
        # Corrigido para coluna 13 (Paletização) conforme sua estrutura
        ids = df.iloc[:, 13].dropna().unique().tolist()
        ids_limpos = sorted([str(x).strip() for x in ids if str(x).lower() != 'nan' and str(x).strip() != ''])
        
        return ids_limpos
    except Exception as e:
        print(f"Erro: {e}")
        return []
    
@app.get("/relatorio-geral")
async def relatorio_geral():
    try:
        if not os.path.exists(PATH_RELATORIO):
            return []
        
        # Lê o Excel tratando tudo como string para evitar NaNs numéricos
        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl', dtype=str)
        
        # Substitui qualquer valor nulo ou 'nan' por hifen antes de converter para dicionário
        df = df.fillna("-")
        
        # Limpa os nomes das colunas
        df.columns = [str(c).strip() for c in df.columns]

        lista_final = []
        for _, row in df.iterrows():
            # Função interna para buscar valor com fallback total
            def get_val(nome, index):
                val = row.get(nome)
                # Se não achar por nome ou se for um 'nan' string, tenta por índice
                if pd.isna(val) or str(val).lower() == 'nan' or val is None:
                    if index < len(row):
                        val = row.iloc[index]
                
                # Limpeza final: se ainda for nan, vira hifen
                final_val = str(val).strip()
                if final_val.lower() in ["nan", "none", "nat", "null", ""]:
                    return "-"
                return final_val

            item = {
                "DataHora": get_val("DataHora", 6),
                "Carga": get_val("Carga", 7),
                "Codigo": get_val("Codigo", 0),
                "Remessa": get_val("Remessa", 2),
                "NF": get_val("NF", 9),
                "Cliente": get_val("Cliente", 8),
                "Paletizacao": get_val("Paletizacao", 13),
                "Endereco": get_val("Endereco", 14),
                "Doca": get_val("Doca", 16),
                "Expedicao": get_val("Expedicao", 15)
            }
            lista_final.append(item)

        return lista_final

    except Exception as e:
        print(f"ERRO CRÍTICO NO RELATÓRIO: {e}")
        # Retorna erro amigável em vez de travar a aplicação
        raise HTTPException(status_code=500, detail="Erro ao processar dados do Excel")
    
@app.get("/dashboard-metrics")
async def dashboard_metrics(
    data: Optional[str] = None, 
    carga: Optional[str] = None, 
    regiao: Optional[str] = None, 
    cliente: Optional[str] = None
):
    try:
        if not os.path.exists(PATH_RELATORIO):
            return {
                "total_bipagens": 0, "total_cargas": 0, "total_clientes": 0,
                "producao_diaria": {}, "producao_hora": {}, "top_cargas": {},
                "top_clientes": {}, "por_regiao": {}
            }

        df = pd.read_excel(PATH_RELATORIO, engine='openpyxl')

        # 1. Tratamento de DataHora
        df['DataHora'] = pd.to_datetime(df['DataHora'], errors='coerce')
        df = df.dropna(subset=['DataHora'])
        
        # Criar colunas auxiliares para facilitar o filtro
        df['DataFiltro'] = df['DataHora'].dt.strftime('%Y-%m-%d') # Formato do input type="date"

        # --- APLICAÇÃO DOS FILTROS DINÂMICOS ---
        if data:
            df = df[df['DataFiltro'] == data]
        if carga:
            df = df[df['Carga'].astype(str) == str(carga)]
        if regiao:
            df = df[df['Regiao'] == regiao]
        if cliente:
            df = df[df['Cliente'] == cliente]

        # 2. Produção por Dia
        df_diario = df.copy()
        df_diario['DataCurta'] = df_diario['DataHora'].dt.strftime('%d/%m')
        producao_diaria = df_diario.groupby('DataCurta').size().to_dict()

        # 3. Produção por Hora
        df_hora = df.copy()
        df_hora['HoraApenas'] = df_hora['DataHora'].dt.hour
        contagem_hora = df_hora.groupby('HoraApenas').size().sort_index().to_dict()
        producao_hora = {f"{int(h):02d}h": int(v) for h, v in contagem_hora.items()}

        # 4. Top 5 Cargas
        top_cargas = df['Carga'].value_counts().head(5).to_dict()

        # 5. NOVO: Top 10 Clientes (Volume)
        top_clientes = df['Cliente'].value_counts().head(10).to_dict() if 'Cliente' in df.columns else {}

        # 6. Distribuição por Região
        coluna_regiao = 'Regiao' if 'Regiao' in df.columns else 'Cidade'
        por_regiao = df[coluna_regiao].value_counts().head(5).to_dict()

        return {
            "total_bipagens": len(df),
            "total_cargas": int(df['Carga'].nunique()),
            "total_clientes": int(df['Cliente'].nunique() if 'Cliente' in df.columns else 0),
            "producao_diaria": producao_diaria,
            "producao_hora": producao_hora,
            "top_cargas": top_cargas,
            "top_clientes": top_clientes, # Enviando para o novo gráfico
            "por_regiao": por_regiao
        }

    except Exception as e:
        print(f"Erro no Dashboard: {e}")
        return {"erro": str(e)}

# ROTA ADICIONAL: Para carregar os nomes nos Selects do HTML automaticamente
@app.get("/get-filters")
async def get_filters():
    try:
        if not os.path.exists(PATH_RELATORIO): return {}
        df = pd.read_excel(PATH_RELATORIO)
        return {
            "cargas": sorted(df['Carga'].dropna().unique().astype(str).tolist()),
            "regioes": sorted(df['Regiao'].dropna().unique().tolist()),
            "clientes": sorted(df['Cliente'].dropna().unique().tolist())
        }
    except: return {}

# Caminho para salvar os dados (Simulando um Banco de Dados)
DB_FILE = "db_ctos.json"

def carregar_dados():
    if not os.path.exists(DB_FILE):
        return []
    with open(DB_FILE, "r") as f:
        return json.load(f)

def salvar_dados(dados):
    with open(DB_FILE, "w") as f:
        json.dump(dados, f, indent=4)

# --- ROTAS ---

@app.get("/get-next-cto")
async def get_next_cto():
    dados = carregar_dados()
    proximo_numero = len(dados) + 1
    # Formato: CTO-2026-0001
    return {"next_id": f"CTO-2026-{proximo_numero:04d}"}

@app.post("/abrir-cto")
async def abrir_cto(
    id_cto: str = Form(...),
    tipo_ocorrencia: str = Form(...),
    tipo_processo: str = Form(...),
    descricao: str = Form(...),
    arquivos: List[UploadFile] = File(None)
):
    try:
        # 1. Configurar o E-mail
        msg = EmailMessage()
        msg['Subject'] = f"🚨 NOTIFICAÇÃO DE OCORRÊNCIA - {id_cto} | {tipo_ocorrencia}"
        msg['From'] = "micael.dutra@ache.com.br"
        msg['To'] = "nilcelena.reis@ache.com.br"

        corpo_email = f"""
        Prezada Nilcelena,

        Informamos que uma nova ocorrência operacional foi registrada no sistema CD & LM.
        Abaixo seguem os detalhes para análise e tratativa:

        DETALHES DO REGISTRO:
        --------------------------------------------------
        NÚMERO DO CTO:      {id_cto}
        TIPO DE OCORRÊNCIA: {tipo_ocorrencia}
        TIPO DE PROCESSO:   {tipo_processo}
        
        DESCRIÇÃO DOS FATOS:
        {descricao}
        --------------------------------------------------

        As evidências fotográficas foram anexadas a este e-mail.

        Atenciosamente,
        Sistema de Gestão - CD & LM Logistics
        """
        msg.set_content(corpo_email)

        # 2. Processar e Anexar Fotos
        if arquivos:
            for foto in arquivos:
                conteudo = await foto.read()
                # Detectar subtipo (jpg, png, etc)
                ext = foto.filename.split('.')[-1].lower()
                subtype = 'jpeg' if ext in ['jpg', 'jpeg'] else ext
                
                msg.add_attachment(
                    conteudo,
                    maintype='image',
                    subtype=subtype,
                    filename=foto.filename
                )
                # Opcional: Salvar localmente se desejar manter backup no servidor
                # with open(f"log_{id_cto}_{foto.filename}", "wb") as f:
                #     f.write(conteudo)

        # 3. Envio via Servidor SMTP do Outlook (Office 365)
        # O Outlook exige STARTTLS na porta 587
        with smtplib.SMTP('smtp.office365.com', 587) as smtp:
            smtp.starttls()  # Ativa a criptografia necessária pelo Outlook
            smtp.login("micael.dutra@ache.com.br", "Ache@2118")
            smtp.send_message(msg)

        return {"status": "success", "id": id_cto}

    except Exception as e:
        print(f"Erro ao processar CTO: {str(e)}")
        # Retorna o erro detalhado para ajudar no debug se o login falhar
        raise HTTPException(status_code=500, detail=f"Erro ao enviar notificação: {str(e)}")

def enviar_email_notificacao(id, ocorrencia, processo, desc, fotos):
    # --- CONFIGURAÇÕES DE E-MAIL (Ajuste aqui) ---
    EMAIL_REMETENTE = "seu_email@gmail.com"
    SENHA_APP = "sua_senha_de_app_aqui" # Não é a senha normal, é a senha de app do Google
    EMAIL_DESTINATARIO = "responsavel_cto@empresa.com"

    msg = EmailMessage()
    msg['Subject'] = f"🚨 NOVO CTO ABERTO: {id} ({ocorrencia})"
    msg['From'] = EMAIL_REMETENTE
    msg['To'] = EMAIL_DESTINATARIO

    corpo = f"""
    Olá,
    Um novo CTO foi registrado no sistema CD & LM.

    DETALHES:
    --------------------------------------
    ID DO CTO: {id}
    TIPO DE OCORRÊNCIA: {ocorrencia}
    TIPO DE PROCESSO: {processo}
    
    DESCRIÇÃO:
    {desc}
    --------------------------------------
    Sistema de Gestão Operacional - CD & LM
    """
    msg.set_content(corpo)

    # Anexar Fotos
    if fotos:
        for foto in fotos:
            conteudo = foto.file.read()
            msg.add_attachment(
                conteudo, 
                maintype='image', 
                subtype='jpeg', 
                filename=foto.filename
            )

    # Envio via SMTP (Exemplo Gmail)
    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
        smtp.login(EMAIL_REMETENTE, SENHA_APP)
        smtp.send_message(msg)
        
if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)